from django.http import JsonResponse
from django.utils.timezone import make_aware, localtime
from django.db.models import Q,Max,Sum,Count,Case,When,BooleanField,Value,OuterRef,Subquery

from datetime import datetime
import pandas as pd
import io
from decimal import Decimal
import os
from openpyxl.styles import Font, PatternFill

from trade.models import *
from common.models import Status
from partners.models import Partner
from leasing.models import Installment

def is_valid_trade_account_data(data):
    if not data.get('account_id') or not data.get('trade_account'):
        return False, JsonResponse({'message': 'Fill required fields.','status':'error'}, status=400)
    return True, None

def import_leases(self, df_json):
        df = pd.read_json(io.StringIO(df_json), orient='records')
        
        required_columns = []
        empty_rows = df[required_columns].isnull().any(axis=1)
        if empty_rows.any():
            self.process.status = "rejected"
            self.process.save()
            self.process.delete()
            return

        self.process.status = "in_progress"
        self.process.items_count = len(df)
        self.process.save()
        
        previous_progress = 0
        for index,row in df.iterrows():
            current_progress = ((index + 1)/len(df))*100

            if current_progress - previous_progress >= 5:
                self.process.progress = int(current_progress)
                self.process.save()
                previous_progress = current_progress
            
           

        self.process.progress = 100
        self.process.status = "completed"
        self.process.save()

def export_trade_transactions_for_customer(self):
    objs = TradeTransaction.objects.select_related("company","lease","currency").filter(
        Q(lease__uuid=self.params.get('lease')) &
        ~Q(delete_status__in=['2']) &
        ~Q(description__icontains='Kira Ödemeleri')
    ).exclude(delete_status__in=['2']).order_by('posting_group_id','due_date','record_date','trade_transaction_id')

    installments = Installment.objects.select_related("company","lease__currency").filter(
        Q(lease__uuid=self.params.get('lease')) &
        Q(payment_date__lte=localtime().date())
    ).order_by('sequency','payment_date')

    self.process.status = "in_progress"
    self.process.items_count = len(objs)
    self.process.save()

    result = []

    transaction_sequency = 1
    for trade_transaction in objs:
        result.append({
            "uuid": trade_transaction.uuid,
            "transaction_type": "trade_transaction",
            "amount_type": trade_transaction.amount_type,
            "date_obj": localtime(trade_transaction.due_date).date(),
            "date": localtime(trade_transaction.due_date).date().strftime('%d.%m.%Y'),
            "posting_group_id": trade_transaction.posting_group_id,
            "posting_group_name": trade_transaction.posting_group_name,
            "description": trade_transaction.description,
            "amount": trade_transaction.amount,
            "currency": trade_transaction.currency.code if trade_transaction.currency else None,
            "overdue_days": 0,
            "applied_status": "",
            "sequency": transaction_sequency if trade_transaction.amount_type == '0' else 0
        })
        transaction_sequency += 1

    for installment in installments:
        if installment.type == "2":
            description = "Peşinat Vadesi"
        elif installment.type == "5":
            description = "Devir Bedeli Vadesi"
        else:
            description = f"{installment.sequency}. Kira Taksiti Vadesi"

        result.append({
            "uuid": installment.uuid,
            "transaction_type": "installment",
            "amount_type": "1",
            "date_obj": installment.payment_date,
            "date": installment.payment_date.strftime('%d.%m.%Y'),
            "posting_group_id": "1",
            "posting_group_name": "Kira",
            "description": description,
            "amount": installment.amount,
            "currency": installment.lease.currency.code if installment.lease.currency else None,
            "overdue_days": (localtime().date() - installment.payment_date).days,
            "applied_status": "Ödenmedi",
            "sequency": 0
        })

    result.sort(key=lambda x: (x['posting_group_id'], x['date_obj'], -int(x['amount_type'])))

    for item in result:
        # if item["date_obj"] > localtime().date():
        #     balance = {
        #         "balance": "",
        #     }
        #     item["balances"] = balance
        #     continue
        objs = result
        prev_balance = 0
        group = ""
        for o in objs:
            if group != "" and group != o["posting_group_id"]:
                prev_balance = 0
            current_amount = o["amount"] if o["amount_type"] == '1' else -o["amount"]
            prev_balance += current_amount
            if o["uuid"] == item["uuid"]:
                balance = {
                    "balance": prev_balance,
                }
                break
            group = o["posting_group_id"]
        item["balances"] = balance

    remaining_amount = Decimal('0.00')
    payment_sequency = 1
    for index, item in enumerate(filter(lambda x: x["transaction_type"] == "installment", result)):
        skip = False

        payments = list(filter(lambda x: x["transaction_type"] == "trade_transaction" and x["posting_group_id"] == item["posting_group_id"] and x["amount_type"] == '0' and x["sequency"] >= payment_sequency, result))
        item_remaining = item["amount"]
        
        for payment in payments:
            remaining_amount += payment["amount"]

            if remaining_amount >= item_remaining:
                remaining_amount -= item_remaining
                item["overdue_days"] = (payment["date_obj"] - item["date_obj"]).days
                item["applied_status"] = "Ödendi"
                skip = True
                payment_sequency = payment["sequency"] + 1
                break

        if skip:
            continue
    
    data = {
        "Tarih": [],
        "Açıklama": [],
        "Tutar": [],
        "PB": [],
        "Bakiye": [],
        "Gecikme": [],
        "Durum": [],
    }

    previous_progress = 0
    metin = ""
    for index,obj in enumerate(result):
        current_progress = ((index + 1)/len(result))*100

        if current_progress - previous_progress >= 5:
            self.process.progress = int(current_progress)
            self.process.save()
            previous_progress = current_progress

        data["Tarih"].append(obj["date"])
        data["Açıklama"].append(obj["description"])
        data["Tutar"].append(f"-{obj['amount']}" if obj["amount_type"] == '0' else f"{obj['amount']}")
        data["PB"].append(obj["currency"])
        data["Bakiye"].append(obj["balances"]["balance"])
        data["Gecikme"].append(f"-{obj['overdue_days']}" if obj["amount_type"] == '1' else "")
        data["Durum"].append(obj["applied_status"])

    df = pd.DataFrame(data)
    df = df.drop_duplicates()

    numeric_columns = [
        "Bakiye",
        "Tutar"
    ]

    for col in numeric_columns:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    
    base_path = os.path.join(os.getcwd(), "media", "docs", str(self.user.user_companies.filter(is_active=True).first().company.uuid), "trade", "trade_transactions", "documents")
    if not os.path.exists(base_path):
            os.makedirs(base_path)



    excel_dosyasi_adi = f"{base_path}/{datetime.today().strftime('%d-%m-%Y')}-odeme-ekstresi.xlsx"
    with pd.ExcelWriter(excel_dosyasi_adi, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Sayfa', index=False)

            # Workbook'u al
            workbook = writer.book
            worksheet = writer.sheets['Sayfa']

            # Kolon isimlerine göre format uygula
            for idx, col in enumerate(df.columns, 1):  # enumerate 1'den başlıyor
                if col in numeric_columns:
                    for cell in worksheet.iter_cols(min_col=idx, max_col=idx, min_row=2):
                        for c in cell:
                            c.number_format = '#,##0.00'   # İstediğin format

            # Hücre rengi değiştirme
            description_col_idx = df.columns.get_loc("Açıklama") + 1
            applied_status_col_idx = df.columns.get_loc("Durum") + 1
            for row_idx, tutar_degeri in enumerate(df["Tutar"], start=2):
                if tutar_degeri and Decimal(str(tutar_degeri).replace(".", "").replace(",", ".")) > 0:
                    cell = worksheet.cell(row=row_idx, column=description_col_idx)
                    cell.font = Font(color="FF0000")
            # for row_idx, value in enumerate(df["Gecikme"], start=2):
            #     if value and int(value) > 0:
            #         cell = worksheet.cell(row=row_idx, column=applied_status_col_idx)
            #         cell.font = Font(color="FF0000")
    


    self.process.progress = 100
    #self.process.status = "completed"
    self.process.save()

